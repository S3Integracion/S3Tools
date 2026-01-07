# -*- coding: utf-8 -*-
"""Formato engine.

Normalizes the first two WebScraper headers in CSV/XLSX files.
"""
import csv
import io
import json
import os
import re
import sys
import traceback
from pathlib import Path

TEMPLATE_FILES = {
    "tiendas": "PlantillaSitemapsTiendas.json",
    "bbvs": "PlantillaSitemapsBBvs.json",
}
DEFAULT_TEMPLATE = "tiendas"

_TEMPLATE_CACHE = {}
_EXPECTED_CACHE = {}

NORMALIZE_RE = re.compile(r"[^a-zA-Z0-9_']")
FIRST_HEADERS = ["web_scraper_order", "web_scraper_start_url"]


def _app_dir():
    return Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))


def _reset_cwd():
    try:
        app_dir = _app_dir()
        try:
            os.chdir(os.path.relpath(str(app_dir)))
        except Exception:
            os.chdir(app_dir)
    except Exception:
        pass


def find_template_path(template_name):
    candidates = [
        Path(template_name),
        Path.cwd() / "Engines" / "Sitemap" / template_name,
        _app_dir() / template_name,
        _app_dir() / "Sitemap" / template_name,
        _app_dir().parent / "Sitemap" / template_name,
        _app_dir().parent / "Engines" / "Sitemap" / template_name,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def load_template(template_key):
    if template_key in _TEMPLATE_CACHE:
        return _TEMPLATE_CACHE[template_key]
    if template_key not in TEMPLATE_FILES:
        raise ValueError(f"Unknown template key: {template_key}")
    template_name = TEMPLATE_FILES[template_key]
    path = find_template_path(template_name)
    if not path:
        raise FileNotFoundError(f"Sitemap template not found: {template_name}")
    template = json.loads(path.read_text(encoding="utf-8"))
    _TEMPLATE_CACHE[template_key] = template
    return template


def expected_headers(template_key):
    if template_key in _EXPECTED_CACHE:
        return _EXPECTED_CACHE[template_key]
    template = load_template(template_key)
    headers = ["web_scraper_order", "web_scraper_start_url"]
    for selector in template.get("selectors", []):
        selector_type = str(selector.get("type") or "").lower()
        if "elementclick" in selector_type:
            continue
        selector_id = selector.get("id")
        if selector_id:
            headers.append(str(selector_id))
    _EXPECTED_CACHE[template_key] = headers
    return headers


def normalize_header(value):
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.replace("-", "_").replace(" ", "_")
    text = NORMALIZE_RE.sub("_", text)
    text = re.sub(r"_+", "_", text)
    return text.lower()


def detect_template(headers):
    normalized_headers = {normalize_header(h) for h in headers if h is not None}
    best = DEFAULT_TEMPLATE
    best_count = -1
    for key in TEMPLATE_FILES:
        expected = expected_headers(key)
        expected_norm = {
            normalize_header(h)
            for h in expected
            if not str(h).lower().startswith("web_scraper")
        }
        count = len(normalized_headers & expected_norm)
        if count > best_count:
            best_count = count
            best = key
    if best_count <= 0:
        return DEFAULT_TEMPLATE
    return best


def resolve_template(choice, headers):
    choice = (choice or "").strip().lower()
    if choice in TEMPLATE_FILES:
        return choice
    return detect_template(headers)

def apply_first_headers(headers):
    updated = list(headers)
    if len(updated) >= 1:
        updated[0] = FIRST_HEADERS[0]
    if len(updated) >= 2:
        updated[1] = FIRST_HEADERS[1]
    return updated


def read_text_with_encoding(path):
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return Path(path).read_text(encoding=enc), enc
        except Exception:
            continue
    return Path(path).read_text(encoding="utf-8", errors="ignore"), "utf-8"


def detect_csv_delimiter(sample):
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";\t,|")
        return dialect.delimiter
    except Exception:
        if ";" in sample and "," not in sample:
            return ";"
        if "\t" in sample:
            return "\t"
        if "|" in sample:
            return "|"
        return ","


def update_csv_headers(path, template_choice):
    text, encoding = read_text_with_encoding(path)
    if not text.strip():
        raise RuntimeError("Empty CSV file")
    delimiter = detect_csv_delimiter(text[:4096])
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows:
        raise RuntimeError("Empty CSV file")
    headers = rows[0]
    template_key = resolve_template(template_choice, headers)
    rows[0] = apply_first_headers(headers)
    with open(path, "w", encoding=encoding, newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        writer.writerows(rows)
    return template_key


def update_xlsx_headers(path, template_choice):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required to edit .xlsx files") from exc
    wb = load_workbook(path)
    try:
        template_key = template_choice
        if (template_choice or "").strip().lower() not in TEMPLATE_FILES:
            template_key = None
            for ws in wb.worksheets:
                if ws.max_column < 1:
                    continue
                headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
                if headers:
                    template_key = detect_template(headers)
                    break
            if not template_key:
                template_key = DEFAULT_TEMPLATE
        for ws in wb.worksheets:
            if ws.max_column < 1:
                continue
            if ws.max_column >= 1:
                ws.cell(row=1, column=1).value = FIRST_HEADERS[0]
            if ws.max_column >= 2:
                ws.cell(row=1, column=2).value = FIRST_HEADERS[1]
        wb.save(path)
    finally:
        wb.close()
    return template_key


def update_headers_in_file(path, template_choice):
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return update_csv_headers(path, template_choice)
    if ext == ".xlsx":
        return update_xlsx_headers(path, template_choice)
    raise RuntimeError("Unsupported file extension. Use .csv or .xlsx.")


def _error(message, tb=None):
    return {"ok": False, "error": message, "traceback": tb or ""}


def handle_process(data):
    _reset_cwd()

    input_files = data.get("input_files") or []
    input_files = [f for f in input_files if f]
    if not input_files:
        return _error("Missing input_files")

    template_choice = (data.get("template") or "").strip().lower()
    template_choice = template_choice if template_choice else "auto"

    updated_files = []
    template_counts = {}

    for fp in input_files:
        if not Path(fp).exists():
            return _error(f"Input file not found: {fp}")
        try:
            template_key = update_headers_in_file(fp, template_choice)
            updated_files.append(fp)
            template_counts[template_key] = template_counts.get(template_key, 0) + 1
        except Exception as exc:
            return _error(f"Failed to update {fp}: {exc}", traceback.format_exc())

    return {
        "ok": True,
        "updated_files": updated_files,
        "template_counts": template_counts,
    }


def handle_request(data):
    action = (data.get("action") or "").strip().lower()
    if action == "process":
        return handle_process(data)
    return _error("Unknown action")


def main():
    raw = sys.stdin.read()
    if not raw.strip():
        resp = _error("No input received")
    else:
        try:
            data = json.loads(raw)
            resp = handle_request(data)
        except Exception as exc:
            resp = _error(str(exc), traceback.format_exc())
    sys.stdout.write(json.dumps(resp, ensure_ascii=True))


if __name__ == "__main__":
    main()
